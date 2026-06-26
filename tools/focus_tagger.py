#!/usr/bin/env python3
"""
ZooDex image tagger
===================

A small desktop tool for framing species photos, checking their on-phone quality,
and recording image copyright.

The list shows every species in the catalogue, whether or not it has a photo yet,
so missing images are easy to find and work through. A marker on each row shows
whether an image exists, whether a focus point is set, and whether a credit is
recorded. A "Missing images only" filter and a search box narrow the list.

For each species you can:
  * click a focus point: the spot that must stay visible when the app crops the
    photo to the square grid tile or the tall species-page hero;
  * preview the grid and hero crops at the actual pixel sizes a phone renders
    them at, with an upscaling verdict, and open them 1:1 to judge sharpness;
  * type the image copyright/attribution (auto-filled from
    tools/reports/image_credits.csv when the downloader has a row);
  * import a new photo (any format), converted to images/<slug>.webp, with the
    slug pre-filled from the selected species;
  * delete the current image from disk.

It writes two files the app reads:
  assets/data/image_focus.json    {"focus":   {slug: [x, y]}}   x,y in 0..1
  assets/data/image_credits.json  {"credits": {slug: "..."}}

Run:   python tools/focus_tagger.py [project_root]
(project_root defaults to the current directory)

Build a double-clickable .exe (optional):
    pip install pyinstaller pillow
    pyinstaller --onefile --windowed tools/focus_tagger.py
The .exe lands in dist/. Run it from the project root (or pass the root when
prompted) so it can find images/ and assets/data/.
"""

import json
import sys
import tempfile
import webbrowser
from pathlib import Path

try:
    from PIL import Image
except ImportError:
    sys.exit("This tool needs Pillow:  pip install pillow")

PLACEHOLDER_PREFIX = "default"
IMG_EXTS = (".webp", ".png", ".jpg", ".jpeg")


# --------------------------------------------------------------------------- #
# Pure helpers (no GUI), unit-testable
# --------------------------------------------------------------------------- #
def project_paths(root: Path):
    return {
        "images": root / "images",
        "focus": root / "assets" / "data" / "image_focus.json",
        "credits": root / "assets" / "data" / "image_credits.json",
        "catalog": root / "assets" / "data" / "species_catalog.json",
        "reports": root / "tools" / "reports",
    }


def list_image_slugs(images_dir: Path):
    """Map slug -> Path for real photos in images/, excluding placeholders.
    Prefers .webp when both .webp and another extension exist for a slug."""
    out = {}
    if not images_dir.is_dir():
        return out
    files = [p for p in images_dir.iterdir()
             if p.suffix.lower() in IMG_EXTS
             and not p.name.lower().startswith(PLACEHOLDER_PREFIX)]
    files.sort(key=lambda p: (p.stem, IMG_EXTS.index(p.suffix.lower())))
    for p in files:
        out.setdefault(p.stem, p)
    return out


def delete_image_files(images_dir: Path, slug: str):
    """Delete every images/<slug>.<ext> on disk. Returns the deleted paths."""
    deleted = []
    for ext in IMG_EXTS:
        p = images_dir / f"{slug}{ext}"
        if p.exists():
            p.unlink()
            deleted.append(p)
    return deleted


def load_map(path: Path, key: str):
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        inner = data.get(key, data) if isinstance(data, dict) else {}
        return dict(inner) if isinstance(inner, dict) else {}
    except Exception:
        return {}


def save_map(path: Path, key: str, mapping: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps({key: mapping}, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )


def load_catalog_names(catalog_path: Path):
    """Ordered dict slug -> common_name for every catalogue entry with a slug."""
    try:
        cat = json.loads(catalog_path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    out = {}
    for s in cat:
        slug = (s.get("slug") or "").strip()
        if slug and slug not in out:
            out[slug] = s.get("common_name", "") or ""
    return out


def compose_credit(row: dict) -> str:
    """Build a copyright string from a credits row. The downloader's
    `attribution` field is already a full human-readable credit; if it is blank,
    fall back to source / licence / url."""
    attr = (row.get("attribution") or "").strip()
    if attr:
        return attr
    src = (row.get("source") or "").strip()
    lic = (row.get("license") or "").strip()
    url = (row.get("source_url") or "").strip()
    bits = [b for b in [src, lic.upper() if lic else "", url] if b]
    return " - ".join(bits)


def _read_csv_credits(path: Path) -> dict:
    import csv
    out = {}
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            slug = (row.get("slug") or "").strip()
            if slug:
                out[slug] = compose_credit({(k or "").lower(): v for k, v in row.items()})
    return out


def _read_xlsx_credits(path: Path) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c is not None else "" for c in next(rows)]
    idx = {name: i for i, name in enumerate(header) if name}
    out = {}
    for r in rows:
        def g(col):
            i = idx.get(col)
            return "" if i is None or i >= len(r) or r[i] is None else str(r[i])
        slug = g("slug").strip()
        if slug:
            out[slug] = compose_credit({h: g(h) for h in idx})
    wb.close()
    return out


def load_credit_suggestions(reports_dir: Path) -> dict:
    """slug -> suggested copyright string, from tools/reports. Prefers
    image_credits.xlsx, else the image_credits.csv the downloader emits. Never
    fatal."""
    xlsx = reports_dir / "image_credits.xlsx"
    csvf = reports_dir / "image_credits.csv"
    if xlsx.exists():
        try:
            return _read_xlsx_credits(xlsx)
        except ImportError:
            print("image_credits.xlsx found but openpyxl is not installed "
                  "(pip install openpyxl); falling back to the .csv.")
        except Exception as e:
            print(f"Could not read image_credits.xlsx ({e}); falling back to .csv.")
    if csvf.exists():
        try:
            return _read_csv_credits(csvf)
        except Exception as e:
            print(f"Could not read image_credits.csv ({e}).")
    return {}


def cover_crop(img: "Image.Image", tw: int, th: int, fx: float, fy: float):
    """Crop `img` to a tw:th box using cover semantics, framed around the focal
    point (fx, fy in 0..1). Mirrors Flutter BoxFit.cover plus Alignment so the
    preview matches the app."""
    iw, ih = img.size
    scale = max(tw / iw, th / ih)
    cw, ch = tw / scale, th / scale
    cx, cy = fx * iw, fy * ih
    left = min(max(cx - cw / 2, 0), iw - cw)
    top = min(max(cy - ch / 2, 0), ih - ch)
    box = (round(left), round(top), round(left + cw), round(top + ch))
    return img.crop(box).resize((tw, th), Image.LANCZOS)


def slugify(name: str) -> str:
    keep = []
    for ch in name.strip().lower():
        if ch.isalnum():
            keep.append(ch)
        elif ch in " -_":
            keep.append("_")
    s = "".join(keep)
    while "__" in s:
        s = s.replace("__", "_")
    return s.strip("_")


def import_image(src: Path, slug: str, images_dir: Path,
                 max_dim: int = 1400, quality: int = 85) -> Path:
    """Copy/convert an arbitrary image into images/<slug>.webp."""
    images_dir.mkdir(parents=True, exist_ok=True)
    im = Image.open(src).convert("RGB")
    w, h = im.size
    if max(w, h) > max_dim:
        s = max_dim / max(w, h)
        im = im.resize((round(w * s), round(h * s)), Image.LANCZOS)
    dest = images_dir / f"{slug}.webp"
    im.save(dest, "WEBP", quality=quality, method=6)
    return dest


# --------------------------------------------------------------------------- #
# Phone display geometry, so previews match the app's on-device pixel sizes.
#   Hero:  full screen width, 400 logical px tall, BoxFit.cover.
#   Grid:  3 square columns, 12px outer padding, 2px gaps (zoodex_screen.dart).
# Adjust the phone profile to model a different device.
# --------------------------------------------------------------------------- #
HERO_HEIGHT = 400
GRID_COLS = 3
GRID_H_PAD = 12
GRID_GAP = 2
PHONE_LOGICAL_W = 412   # representative phone width in dp (e.g. Pixel 7)
PHONE_DPR = 3.0         # representative device pixel ratio (a demanding case)


def phone_targets(logical_w: int = PHONE_LOGICAL_W, dpr: float = PHONE_DPR):
    """Physical pixel size each context is displayed at on a phone."""
    hero = (round(logical_w * dpr), round(HERO_HEIGHT * dpr))
    tile = (logical_w - 2 * GRID_H_PAD - (GRID_COLS - 1) * GRID_GAP) / GRID_COLS
    grid = (round(tile * dpr), round(tile * dpr))
    return {"Hero": hero, "Grid tile": grid}


def upscale_verdict(iw: int, ih: int, tw: int, th: int):
    """How much the source is scaled to fill a tw x th cover crop. Returns
    (message, colour). scale > 1 means upscaling, which costs sharpness."""
    scale = max(tw / iw, th / ih)
    if scale <= 1.0:
        return f"sharp (downscaled {1 / scale:.1f}x)", "#080"
    if scale < 1.4:
        return f"slightly upscaled {scale:.2f}x", "#a60"
    return f"upscaled {scale:.2f}x - soft", "#c00"


# --------------------------------------------------------------------------- #
# GUI
# --------------------------------------------------------------------------- #
def run_gui(root_dir: Path):
    import tkinter as tk
    from tkinter import filedialog, messagebox, simpledialog
    from PIL import ImageTk, ImageDraw

    paths = project_paths(root_dir)
    focus = load_map(paths["focus"], "focus")
    credits = load_map(paths["credits"], "credits")
    names = load_catalog_names(paths["catalog"])
    suggestions = load_credit_suggestions(paths["reports"])
    slug_files = list_image_slugs(paths["images"])

    master = sorted(set(names) | set(slug_files),
                    key=lambda s: (names.get(s, "").lower(), s))
    visible = list(master)

    app = tk.Tk()
    app.title(f"ZooDex image tagger - {root_dir}")
    app.geometry("1100x680")

    state = {"slug": None, "img": None, "tkmain": None,
             "main_box": (0, 0, 1, 1), "preview_refs": []}

    left = tk.Frame(app, width=300)
    left.pack(side="left", fill="y")
    filt = tk.Frame(left)
    filt.pack(fill="x", padx=6, pady=(6, 0))
    search_var = tk.StringVar()
    tk.Label(filt, text="Search").pack(anchor="w")
    tk.Entry(filt, textvariable=search_var).pack(fill="x")
    missing_var = tk.BooleanVar(value=False)
    tk.Checkbutton(filt, text="Missing images only",
                   variable=missing_var).pack(anchor="w", pady=(4, 0))
    count_lbl = tk.Label(left, text="", anchor="w", fg="#444", font=("", 8))
    count_lbl.pack(anchor="w", padx=6)
    listbox = tk.Listbox(left, width=42)
    listbox.pack(fill="y", expand=True, padx=6, pady=6)

    center = tk.Frame(app)
    center.pack(side="left", fill="both", expand=True)
    canvas = tk.Canvas(center, bg="#222", highlightthickness=0)
    canvas.pack(fill="both", expand=True, padx=8, pady=8)
    tk.Label(center, text="Click the photo to set the focus point").pack()

    right = tk.Frame(app, width=340)
    right.pack(side="left", fill="y")
    targets = phone_targets()
    tk.Label(right,
             text=f"On-phone preview ({PHONE_LOGICAL_W}dp wide, DPR {PHONE_DPR:g})"
             ).pack(anchor="w", padx=8, pady=(8, 0))
    prev_row = tk.Frame(right)
    prev_row.pack(padx=8, pady=4)
    disp_cap = {"Hero": 190, "Grid tile": 130}
    preview_widgets = []
    for ctx in ("Hero", "Grid tile"):
        tw, th = targets[ctx]
        cap = disp_cap[ctx]
        ds = min(cap / tw, cap / th, 1.0)
        dw, dh = max(1, round(tw * ds)), max(1, round(th * ds))
        col = tk.Frame(prev_row)
        col.pack(side="left", padx=6, anchor="n")
        c = tk.Canvas(col, width=dw, height=dh, bg="#333", highlightthickness=0)
        c.pack()
        tk.Label(col, text=f"{ctx}\n{tw}x{th}px", font=("", 8),
                 justify="center").pack()
        v = tk.Label(col, text="", font=("", 8), wraplength=cap + 24,
                     justify="center")
        v.pack()
        preview_widgets.append({"ctx": ctx, "canvas": c, "vlabel": v,
                                "tw": tw, "th": th, "dw": dw, "dh": dh})
    src_lbl = tk.Label(right, text="", font=("", 8), fg="#666", anchor="w")
    src_lbl.pack(anchor="w", padx=8, pady=(2, 0))

    tk.Label(right, text="Copyright / attribution").pack(anchor="w", padx=8, pady=(10, 0))
    credit_var = tk.StringVar()
    tk.Entry(right, textvariable=credit_var, width=40).pack(padx=8, pady=4, fill="x")
    status = tk.Label(right, text="", fg="#080", anchor="w", justify="left",
                      wraplength=320)
    status.pack(anchor="w", padx=8, pady=6, fill="x")

    def has_img(slug):
        return slug in slug_files

    def list_label(slug):
        img_mark = "[img]" if has_img(slug) else "[   ]"
        foc = "*" if slug in focus else " "
        cr = "(c)" if credits.get(slug) else ("." if slug in suggestions else "  ")
        nm = names.get(slug, "")
        return f"{img_mark} {foc}{cr} {slug}" + (f"  ({nm})" if nm else "")

    def recompute_visible():
        nonlocal visible
        q = search_var.get().strip().lower()
        only_missing = missing_var.get()
        visible = [s for s in master
                   if (not only_missing or not has_img(s))
                   and (not q or q in s.lower() or q in names.get(s, "").lower())]

    def update_count():
        total = len(master)
        with_img = sum(1 for s in master if has_img(s))
        count_lbl.config(
            text=f"{with_img}/{total} have images - {total - with_img} missing"
                 f" - showing {len(visible)}")

    def refresh_list(select=None):
        recompute_visible()
        listbox.delete(0, "end")
        for sl in visible:
            listbox.insert("end", list_label(sl))
        if select in visible:
            i = visible.index(select)
            listbox.selection_clear(0, "end")
            listbox.selection_set(i)
            listbox.see(i)
        update_count()

    def draw_previews():
        state["preview_refs"].clear()
        img = state["img"]
        if img is None:
            src_lbl.config(text="")
            for w in preview_widgets:
                w["canvas"].delete("all")
                w["vlabel"].config(text="")
            return
        iw, ih = img.size
        src_lbl.config(text=f"Source image: {iw} x {ih} px")
        fx, fy = focus.get(state["slug"], [0.5, 0.5])
        for w in preview_widgets:
            cv = w["canvas"]
            cv.delete("all")
            crop = cover_crop(img, w["tw"], w["th"], fx, fy)
            disp = crop.resize((w["dw"], w["dh"]), Image.LANCZOS)
            tkimg = ImageTk.PhotoImage(disp)
            cv.create_image(w["dw"] // 2, w["dh"] // 2, image=tkimg)
            state["preview_refs"].append(tkimg)
            msg, col = upscale_verdict(iw, ih, w["tw"], w["th"])
            w["vlabel"].config(text=msg, fg=col)

    def inspect_full():
        """Save the crops at true phone pixels and open them in the OS viewer,
        where they can be zoomed to judge sharpness."""
        img = state["img"]
        if img is None:
            return
        fx, fy = focus.get(state["slug"], [0.5, 0.5])
        tmp = Path(tempfile.gettempdir())
        opened = 0
        for ctx in ("Hero", "Grid tile"):
            tw, th = phone_targets()[ctx]
            crop = cover_crop(img, tw, th, fx, fy)
            tag = ctx.replace(" ", "_").lower()
            out = tmp / f"zoodex_{state['slug']}_{tag}_{tw}x{th}.png"
            crop.save(out)
            webbrowser.open(out.as_uri())
            opened += 1
        status.config(text=f"Opened {opened} crop(s) at 1:1 in your image viewer.")

    def draw_main():
        img = state["img"]
        canvas.delete("all")
        cw = max(canvas.winfo_width(), 50)
        ch = max(canvas.winfo_height(), 50)
        if img is None:
            canvas.create_text(
                cw // 2, ch // 2, fill="#999", width=cw - 40, justify="center",
                text="No image for this species.\nUse Import to add one.")
            return
        iw, ih = img.size
        scale = min(cw / iw, ch / ih)
        dw, dh = int(iw * scale), int(ih * scale)
        ox, oy = (cw - dw) // 2, (ch - dh) // 2
        disp = img.convert("RGBA").resize((dw, dh), Image.LANCZOS)
        fx, fy = focus.get(state["slug"], [0.5, 0.5])
        d = ImageDraw.Draw(disp)
        px, py = int(fx * dw), int(fy * dh)
        d.line([(px - 12, py), (px + 12, py)], fill=(255, 80, 80, 255), width=2)
        d.line([(px, py - 12), (px, py + 12)], fill=(255, 80, 80, 255), width=2)
        d.ellipse([px - 7, py - 7, px + 7, py + 7], outline=(255, 80, 80, 255), width=2)
        tkimg = ImageTk.PhotoImage(disp)
        state["tkmain"] = tkimg
        state["main_box"] = (ox, oy, dw, dh)
        canvas.create_image(ox, oy, image=tkimg, anchor="nw")

    def update_buttons():
        st = "normal" if has_img(state["slug"]) else "disabled"
        del_btn.config(state=st)
        inspect_btn.config(state=st)

    def load_slug(slug):
        state["slug"] = slug
        if has_img(slug):
            try:
                state["img"] = Image.open(slug_files[slug]).convert("RGB")
            except Exception as e:
                state["img"] = None
                messagebox.showerror("Open failed", f"{slug}: {e}")
        else:
            state["img"] = None
        saved = credits.get(slug, "")
        if saved:
            credit_var.set(saved)
            status.config(text="")
        elif slug in suggestions:
            credit_var.set(suggestions[slug])
            status.config(text="Copyright pre-filled from image_credits.csv. "
                               "Edit, then Save to keep.")
        else:
            credit_var.set("")
            status.config(text="" if has_img(slug)
                          else "No image yet. Import one to add a photo.")
        draw_main()
        draw_previews()
        update_buttons()

    def commit_credit():
        if state["slug"] is not None:
            c = credit_var.get().strip()
            if c:
                credits[state["slug"]] = c
            else:
                credits.pop(state["slug"], None)

    def on_canvas_click(ev):
        if state["img"] is None:
            return
        ox, oy, dw, dh = state["main_box"]
        if dw == 0 or dh == 0:
            return
        fx = min(max((ev.x - ox) / dw, 0.0), 1.0)
        fy = min(max((ev.y - oy) / dh, 0.0), 1.0)
        focus[state["slug"]] = [round(fx, 4), round(fy, 4)]
        draw_main()
        draw_previews()
        refresh_list(select=state["slug"])

    def on_select(_ev=None):
        sel = listbox.curselection()
        if not sel:
            return
        commit_credit()
        load_slug(visible[sel[0]])

    def center_focus():
        if state["slug"] and state["img"] is not None:
            focus[state["slug"]] = [0.5, 0.5]
            draw_main()
            draw_previews()
            refresh_list(select=state["slug"])

    def save_all():
        commit_credit()
        save_map(paths["focus"], "focus", focus)
        save_map(paths["credits"], "credits", credits)
        status.config(text=f"Saved {len(focus)} focus point(s), "
                           f"{len(credits)} credit(s).")
        refresh_list(select=state["slug"])

    def import_new():
        src = filedialog.askopenfilename(
            title="Choose an image to import",
            filetypes=[("Images", "*.png *.jpg *.jpeg *.webp *.bmp *.gif")])
        if not src:
            return
        default = state["slug"] or slugify(Path(src).stem)
        slug = simpledialog.askstring(
            "Slug", "Species slug (matches species_catalog 'slug'):",
            initialvalue=default)
        if not slug:
            return
        slug = slugify(slug)
        try:
            import_image(Path(src), slug, paths["images"])
        except Exception as e:
            messagebox.showerror("Import failed", str(e))
            return
        nonlocal slug_files, master
        slug_files = list_image_slugs(paths["images"])
        if slug not in master:
            master = sorted(set(master) | {slug},
                            key=lambda s: (names.get(s, "").lower(), s))
        refresh_list(select=slug)
        load_slug(slug)
        status.config(text=f"Imported {slug}.webp. Now set its focus point.")

    def delete_current():
        slug = state["slug"]
        if not slug or not has_img(slug):
            return
        if not messagebox.askyesno(
                "Delete image",
                f"Delete the image file for '{slug}'?\n\n"
                f"This removes images/{slug}.* from disk and clears its focus "
                f"point. The species stays in the list so a new image can be "
                f"added later."):
            return
        nonlocal slug_files
        try:
            delete_image_files(paths["images"], slug)
        except Exception as e:
            messagebox.showerror("Delete failed", str(e))
            return
        focus.pop(slug, None)
        slug_files = list_image_slugs(paths["images"])
        refresh_list(select=slug)
        load_slug(slug)
        status.config(text=f"Deleted the image for {slug}. Save to persist the "
                           f"cleared focus point.")

    def refresh_images():
        nonlocal slug_files, master, suggestions
        commit_credit()  # keep any unsaved credit edit
        slug_files = list_image_slugs(paths["images"])
        suggestions = load_credit_suggestions(paths["reports"])
        master = sorted(set(names) | set(slug_files),
                        key=lambda s: (names.get(s, "").lower(), s))
        cur = state["slug"]
        refresh_list(select=cur)
        if cur:
            load_slug(cur)
        with_img = sum(1 for s in master if has_img(s))
        status.config(text=f"Refreshed: {with_img}/{len(master)} species have "
                           f"images.")

    btns = tk.Frame(right)
    btns.pack(side="bottom", fill="x", padx=8, pady=10)
    tk.Button(btns, text="Save", command=save_all, width=8).pack(side="left")
    tk.Button(btns, text="Centre", command=center_focus).pack(side="left", padx=4)
    tk.Button(btns, text="Import...", command=import_new).pack(side="left", padx=4)
    tk.Button(btns, text="Refresh", command=refresh_images).pack(side="left", padx=4)
    inspect_btn = tk.Button(btns, text="1:1", command=inspect_full)
    inspect_btn.pack(side="left", padx=4)
    del_btn = tk.Button(btns, text="Delete", command=delete_current, fg="#a00")
    del_btn.pack(side="left", padx=4)

    listbox.bind("<<ListboxSelect>>", on_select)
    canvas.bind("<Button-1>", on_canvas_click)
    canvas.bind("<Configure>", lambda e: draw_main())
    search_var.trace_add("write", lambda *_: refresh_list(select=state["slug"]))
    missing_var.trace_add("write", lambda *_: refresh_list(select=state["slug"]))

    refresh_list()
    if visible:
        listbox.selection_set(0)
        load_slug(visible[0])
    else:
        status.config(text="No species found. Check that species_catalog.json "
                           "and images/ are present.")

    app.mainloop()


def main():
    root = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else Path.cwd()
    if not (root / "images").exists() and not (root / "assets").exists():
        print(f"Warning: {root} does not look like the project root "
              f"(no images/ or assets/). Continuing anyway.")
    run_gui(root)


if __name__ == "__main__":
    main()
